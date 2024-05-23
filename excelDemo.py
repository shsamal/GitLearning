import openpyxl

work_book = openpyxl.load_workbook("/Users/shailajasamal/Documents/Selenium_data/Selenium_E2E_pytestFramework/load_data.xlsx")
# sheet = work_book.active
sheet2 = work_book.get_sheet_by_name("Sheet 2")
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# print(sheet['A3'].value)
# sheet.cell(row=2, column=2).value = "Shailaja"
# print(sheet.cell(row=2, column=2).value)
# print(sheet.max_row)
# print(sheet.max_column)
Dict = {}
lst = []
# whole data set: test cases
for i in range(2, sheet2.max_row+1):
    for j in range(2, sheet2.max_column+1):
        # print(sheet2.cell(row=i, column=j).value)
        Dict[sheet2.cell(row=1, column=j).value] = sheet2.cell(row=i, column=j).value
        # lst.append(Dict)
    print(Dict)
    lst.append(Dict.copy()) # use dict.copy() to append a copy of the dict, it works !!
print(lst)

# Only 1 test case
for i in range(1, sheet2.max_row+1):
    if sheet2.cell(row=i, column=1).value == "TC2":
        for j in range(2, sheet2.max_column+1):
            print(sheet2.cell(row=i, column=j).value)


