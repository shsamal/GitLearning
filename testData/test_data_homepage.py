import openpyxl


class TestData:

    homepage_data = [{"firstname": "Shailaja", "email": "shailajasamal@gmail.com",
                      "password": "Qwerty$#21", "gender": "Female", "lastname": " Samal"},
                     {"firstname": "Rahul", "email": "rahulshetty@gmail.com",
                      "password": "Qwerty$#22", "gender": "Male", "lastname": " Shetty"}
                     ]

    @staticmethod
    def get_testData():
        work_book = openpyxl.load_workbook(
            "/Users/shailajasamal/Documents/Selenium_data/Selenium_E2E_pytestFramework/load_data.xlsx")
        sheet2 = work_book.get_sheet_by_name("Sheet 2")
        Dict = {}
        lst = []
        for i in range(2, sheet2.max_row + 1):
            for j in range(2, sheet2.max_column + 1):
                Dict[sheet2.cell(row=1, column=j).value] = sheet2.cell(row=i, column=j).value
            print(Dict)
            lst.append(Dict.copy())  # use dict.copy() to append a copy of the dict, it works !!
        print(lst)
        return lst



